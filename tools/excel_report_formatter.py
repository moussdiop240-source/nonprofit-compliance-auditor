"""
Excel Report Formatter — complements formatting_tools.py (PDF).
Exports the full audit result to a multi-sheet .xlsx workbook:
  Sheet 1 — Summary
  Sheet 2 — Extracted Line Items
  Sheet 3 — Compliance Decisions
  Sheet 4 — Human Review Decisions (if any)
Requires openpyxl (already in requirements.txt).
"""
import io
import logging
from datetime import datetime
from typing import Any

logger = logging.getLogger(__name__)

# Column widths (characters)
_COL_WIDTHS = {
    "line_number":       8,
    "description":       45,
    "amount":            14,
    "category":          16,
    "vendor":            22,
    "date":              14,
    "status":            24,
    "regulation_cited":  22,
    "reasoning":         50,
    "confidence_score":  18,
    "flagged_reason":    35,
    "human_decision":    20,
    "human_review_note": 40,
    "reviewed_at":       22,
}


def generate_excel_report(audit_state: dict) -> bytes:
    """
    Convert a completed audit state dict to an .xlsx binary.

    Args:
        audit_state: The final AuditState dict returned by the graph.

    Returns:
        Raw .xlsx bytes suitable for st.download_button.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl is required: pip install openpyxl")

    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    # ── Styles ────────────────────────────────────────────────────────────────
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="1E3A5F")
    center_align = Alignment(horizontal="center", vertical="top", wrap_text=False)
    wrap_align   = Alignment(horizontal="left",   vertical="top", wrap_text=True)
    thin_border  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    _STATUS_FILLS = {
        "ALLOWABLE":               PatternFill("solid", fgColor="D1FAE5"),
        "UNALLOWABLE":             PatternFill("solid", fgColor="FEE2E2"),
        "CONDITIONALLY_ALLOWABLE": PatternFill("solid", fgColor="FEF3C7"),
        "REQUIRES_REVIEW":         PatternFill("solid", fgColor="DBEAFE"),
    }

    def _write_header(ws, columns: list[str]) -> None:
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = _COL_WIDTHS.get(col_name, 18)

    def _write_row(ws, row_idx: int, columns: list[str], data: dict,
                   status_col: str = "status") -> None:
        status = data.get(status_col, "")
        row_fill = _STATUS_FILLS.get(status)
        for col_idx, col_name in enumerate(columns, start=1):
            val = data.get(col_name, "")
            if isinstance(val, float):
                val = round(val, 4)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = wrap_align
            cell.border = thin_border
            if row_fill and col_name == status_col:
                cell.fill = row_fill

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws_summary = wb.create_sheet("Summary")
    summary_rows = [
        ("Audit Date",          datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Organization",        audit_state.get("organization_name", "")),
        ("Grant Number",        audit_state.get("grant_number", "")),
        ("Total Line Items",    len(audit_state.get("extracted_line_items", []))),
        ("Total Allowable ($)", round(audit_state.get("total_allowable", 0), 2)),
        ("Total Unallowable ($)", round(audit_state.get("total_unallowable", 0), 2)),
        ("Allowable Items",
         sum(1 for d in audit_state.get("compliance_decisions", []) if d.get("status") == "ALLOWABLE")),
        ("Unallowable Items",
         sum(1 for d in audit_state.get("compliance_decisions", []) if d.get("status") == "UNALLOWABLE")),
        ("Conditionally Allowable",
         sum(1 for d in audit_state.get("compliance_decisions", []) if d.get("status") == "CONDITIONALLY_ALLOWABLE")),
        ("Requires Review",
         sum(1 for d in audit_state.get("compliance_decisions", []) if d.get("status") == "REQUIRES_REVIEW")),
        ("Human Reviewed Items", len(audit_state.get("human_review_decisions", []))),
        ("Standards Applied",   "2 CFR 200 Uniform Guidance | 2026 IRS / GAAP"),
    ]
    ws_summary.column_dimensions["A"].width = 28
    ws_summary.column_dimensions["B"].width = 40
    for r_idx, (label, value) in enumerate(summary_rows, start=1):
        lbl_cell = ws_summary.cell(row=r_idx, column=1, value=label)
        lbl_cell.font = Font(bold=True)
        lbl_cell.border = thin_border
        val_cell = ws_summary.cell(row=r_idx, column=2, value=value)
        val_cell.border = thin_border

    # ── Sheet 2: Line Items ───────────────────────────────────────────────────
    ws_items = wb.create_sheet("Line Items")
    item_cols = ["line_number", "description", "amount", "category", "vendor", "date"]
    _write_header(ws_items, item_cols)
    for r_idx, item in enumerate(audit_state.get("extracted_line_items", []), start=2):
        _write_row(ws_items, r_idx, item_cols, item, status_col="")

    # ── Sheet 3: Compliance Decisions ─────────────────────────────────────────
    ws_decisions = wb.create_sheet("Compliance Decisions")
    dec_cols = [
        "line_number", "description", "amount", "category",
        "status", "regulation_cited", "reasoning",
        "confidence_score", "flagged_reason",
    ]
    _write_header(ws_decisions, dec_cols)
    for r_idx, dec in enumerate(audit_state.get("compliance_decisions", []), start=2):
        _write_row(ws_decisions, r_idx, dec_cols, dec)

    # ── Sheet 4: Human Review (if any) ────────────────────────────────────────
    hr_decisions = audit_state.get("human_review_decisions", [])
    if hr_decisions:
        ws_hr = wb.create_sheet("Human Review")
        hr_cols = [
            "line_number", "description", "amount",
            "human_decision", "human_review_note", "reviewed_at",
        ]
        _write_header(ws_hr, hr_cols)
        for r_idx, hr in enumerate(hr_decisions, start=2):
            _write_row(ws_hr, r_idx, hr_cols, hr, status_col="human_decision")

    # ── Serialise ─────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    logger.info("Excel report generated: %d sheets", len(wb.sheetnames))
    return buf.read()
