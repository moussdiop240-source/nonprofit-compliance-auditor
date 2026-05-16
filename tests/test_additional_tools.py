"""
Tests for the three additional enhancement tools:
  - tools/vectorstore_maintenance.py
  - tools/visualization_tools.py
  - tools/excel_report_formatter.py
"""
import io
import pytest


# ─── vectorstore_maintenance ──────────────────────────────────────────────────

class TestVectorstoreMaintenance:
    def test_get_cfr200_stats_returns_dict(self):
        from tools.vectorstore_maintenance import get_cfr200_stats
        stats = get_cfr200_stats()
        assert isinstance(stats, dict)
        assert "version" in stats
        assert "healthy" in stats
        assert "persist_dir_exists" in stats

    def test_get_grant_store_stats_empty(self, tmp_path):
        from tools.vectorstore_maintenance import get_grant_store_stats
        stats = get_grant_store_stats(base_dir=str(tmp_path))
        assert stats["store_count"] == 0
        assert stats["stores"] == []

    def test_get_grant_store_stats_finds_dirs(self, tmp_path):
        from tools.vectorstore_maintenance import get_grant_store_stats
        (tmp_path / "chroma_grant_abc123").mkdir()
        (tmp_path / "chroma_grant_def456").mkdir()
        (tmp_path / "some_other_dir").mkdir()
        stats = get_grant_store_stats(base_dir=str(tmp_path))
        assert stats["store_count"] == 2

    def test_check_cfr200_health_returns_dict(self):
        from tools.vectorstore_maintenance import check_cfr200_health
        result = check_cfr200_health()
        assert isinstance(result, dict)
        assert "healthy" in result
        assert "latency_ms" in result or "error" in result

    def test_full_maintenance_report_structure(self):
        from tools.vectorstore_maintenance import full_maintenance_report
        report = full_maintenance_report()
        assert "cfr200" in report
        assert "grant_stores" in report
        assert "overall_healthy" in report


# ─── visualization_tools ──────────────────────────────────────────────────────

_SAMPLE_DECISIONS = [
    {"line_number": 1, "description": "Flight", "amount": 450.0,
     "category": "travel", "status": "ALLOWABLE", "confidence_score": 0.85},
    {"line_number": 2, "description": "Alcohol", "amount": 60.0,
     "category": "food", "status": "UNALLOWABLE", "confidence_score": 0.92},
    {"line_number": 3, "description": "Conference fee", "amount": 200.0,
     "category": "professional", "status": "CONDITIONALLY_ALLOWABLE", "confidence_score": 0.55},
    {"line_number": 4, "description": "Laptop", "amount": 1200.0,
     "category": "equipment", "status": "REQUIRES_REVIEW", "confidence_score": 0.40},
]

_SAMPLE_ITEMS = [
    {"line_number": 1, "description": "Flight", "amount": 450.0, "category": "travel"},
    {"line_number": 2, "description": "Alcohol", "amount": 60.0,  "category": "food"},
    {"line_number": 3, "description": "Conference", "amount": 200.0, "category": "professional"},
]


class TestVisualizationTools:
    def test_compliance_breakdown_returns_fig_or_none(self):
        from tools.visualization_tools import compliance_breakdown_chart
        result = compliance_breakdown_chart(_SAMPLE_DECISIONS)
        # Either a plotly Figure or None (plotly not installed)
        assert result is None or hasattr(result, "data")

    def test_compliance_breakdown_empty(self):
        from tools.visualization_tools import compliance_breakdown_chart
        assert compliance_breakdown_chart([]) is None

    def test_expense_by_category_returns_fig_or_none(self):
        from tools.visualization_tools import expense_by_category_chart
        result = expense_by_category_chart(_SAMPLE_ITEMS)
        assert result is None or hasattr(result, "data")

    def test_expense_by_category_empty(self):
        from tools.visualization_tools import expense_by_category_chart
        assert expense_by_category_chart([]) is None

    def test_confidence_distribution_returns_fig_or_none(self):
        from tools.visualization_tools import confidence_distribution_chart
        result = confidence_distribution_chart(_SAMPLE_DECISIONS)
        assert result is None or hasattr(result, "data")

    def test_confidence_distribution_no_scores(self):
        from tools.visualization_tools import confidence_distribution_chart
        assert confidence_distribution_chart([{"status": "ALLOWABLE"}]) is None

    def test_allowable_vs_unallowable_bar(self):
        from tools.visualization_tools import allowable_vs_unallowable_bar
        result = allowable_vs_unallowable_bar(1000.0, 200.0, 150.0)
        assert result is None or hasattr(result, "data")


# ─── excel_report_formatter ───────────────────────────────────────────────────

_SAMPLE_STATE = {
    "organization_name": "Test Nonprofit",
    "grant_number": "G-2024-001",
    "total_allowable": 1000.0,
    "total_unallowable": 200.0,
    "extracted_line_items": [
        {"line_number": 1, "description": "Flight to DC", "amount": 450.0,
         "category": "travel", "vendor": "Delta", "date": "2024-03-15"},
        {"line_number": 2, "description": "Beer at dinner", "amount": 60.0,
         "category": "food", "vendor": "Restaurant", "date": "2024-03-15"},
    ],
    "compliance_decisions": [
        {"line_number": 1, "description": "Flight to DC", "amount": 450.0,
         "category": "travel", "status": "ALLOWABLE",
         "regulation_cited": "2 CFR 200.474", "reasoning": "Economy airfare",
         "confidence_score": 0.85, "flagged_reason": ""},
        {"line_number": 2, "description": "Beer at dinner", "amount": 60.0,
         "category": "food", "status": "UNALLOWABLE",
         "regulation_cited": "2 CFR 200.423", "reasoning": "Alcohol unallowable",
         "confidence_score": 0.92, "flagged_reason": "Alcoholic beverage"},
    ],
    "human_review_decisions": [
        {"line_number": 2, "description": "Beer at dinner", "amount": 60.0,
         "human_decision": "REJECTED", "human_review_note": "Confirmed unallowable",
         "reviewed_at": "2024-03-16T10:00:00"},
    ],
}


class TestExcelReportFormatter:
    def test_generate_excel_returns_bytes(self):
        from tools.excel_report_formatter import generate_excel_report
        result = generate_excel_report(_SAMPLE_STATE)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_excel_is_valid_workbook(self):
        from tools.excel_report_formatter import generate_excel_report
        from openpyxl import load_workbook
        result = generate_excel_report(_SAMPLE_STATE)
        wb = load_workbook(io.BytesIO(result))
        assert "Summary" in wb.sheetnames
        assert "Line Items" in wb.sheetnames
        assert "Compliance Decisions" in wb.sheetnames
        assert "Human Review" in wb.sheetnames

    def test_summary_sheet_has_org_name(self):
        from tools.excel_report_formatter import generate_excel_report
        from openpyxl import load_workbook
        result = generate_excel_report(_SAMPLE_STATE)
        wb = load_workbook(io.BytesIO(result))
        ws = wb["Summary"]
        values = [ws.cell(row=r, column=2).value for r in range(1, 13)]
        assert "Test Nonprofit" in values

    def test_line_items_sheet_row_count(self):
        from tools.excel_report_formatter import generate_excel_report
        from openpyxl import load_workbook
        result = generate_excel_report(_SAMPLE_STATE)
        wb = load_workbook(io.BytesIO(result))
        ws = wb["Line Items"]
        # Header + 2 data rows = 3 rows with content
        assert ws.max_row >= 3

    def test_no_human_review_sheet_when_empty(self):
        from tools.excel_report_formatter import generate_excel_report
        from openpyxl import load_workbook
        state = {**_SAMPLE_STATE, "human_review_decisions": []}
        result = generate_excel_report(state)
        wb = load_workbook(io.BytesIO(result))
        assert "Human Review" not in wb.sheetnames

    def test_empty_state_does_not_crash(self):
        from tools.excel_report_formatter import generate_excel_report
        result = generate_excel_report({
            "organization_name": "", "grant_number": "",
            "total_allowable": 0.0, "total_unallowable": 0.0,
            "extracted_line_items": [], "compliance_decisions": [],
            "human_review_decisions": [],
        })
        assert isinstance(result, bytes)
