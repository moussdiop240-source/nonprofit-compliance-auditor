"""
Generate two pairs of sample grant agreements and expense reports
(PDF + Excel each) for testing the compliance auditor pipeline.

Example 1 — Community Health Initiative (HHS, mostly clean expenses)
Example 2 — Youth Education Foundation  (DoE, mixed allowable/unallowable)
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
)

OUT = os.path.join(os.path.dirname(__file__), "data", "sample_documents")
os.makedirs(OUT, exist_ok=True)

# ─────────────────────────────────────────────────────────────────────────────
# EXAMPLE 1  Community Health Initiative  ·  HHS Award #2026-HHS-101
# ─────────────────────────────────────────────────────────────────────────────

GRANT_1 = {
    "title": "FEDERAL GRANT AGREEMENT",
    "grant_number": "2026-HHS-101",
    "grantee": "Community Health Initiative",
    "agency": "U.S. Department of Health and Human Services",
    "award_amount": "$85,000",
    "period": "January 1, 2026 - December 31, 2026",
    "cfda": "93.600",
    "sections": [
        ("PURPOSE",
         "Funds support a community diabetes prevention program serving low-income "
         "adults in three counties, consistent with 2 CFR 200 (Uniform Guidance)."),

        ("BUDGET SUMMARY",
         "Personnel:           $52,000\n"
         "Fringe Benefits:     $10,400\n"
         "Travel:               $6,000\n"
         "Supplies:             $4,500\n"
         "Contractual:          $8,000\n"
         "Indirect Costs (15%): $4,100\n"
         "Total:               $85,000"),

        ("ALLOWABLE COSTS",
         "All costs must conform to 2 CFR 200.420: necessary, reasonable, and allocable. "
         "Costs must be consistently treated and adequately documented."),

        ("TRAVEL",
         "Travel is allowable per 2 CFR 200.474. Economy/coach class required; "
         "first-class or business-class requires prior written approval. GSA per diem "
         "rates apply for meals ($59/day) and lodging ($150/night). Mileage at the "
         "current IRS standard rate. Receipts required for expenses exceeding $75."),

        ("PERSONNEL",
         "Program Director (50% FTE, $26,000): pre-approved.\n"
         "Health Educator (100% FTE, $26,000): pre-approved.\n"
         "All salaries require time-and-effort certification. Fringe rate: 20%."),

        ("SUPPLIES & EQUIPMENT",
         "Supplies directly used for program activities are allowable. Equipment "
         "purchases over $5,000 require prior written approval. Computing devices "
         "under $5,000 are allowable if essential and not duplicative."),

        ("CONTRACTUAL",
         "Subcontractor services must have documented cost/price analyses. The "
         "nutrition counseling contract ($8,000) is pre-approved. Any new contracts "
         "over $2,500 require prior approval."),

        ("UNALLOWABLE COSTS",
         "The following are explicitly unallowable:\n"
         "- Alcoholic beverages (2 CFR 200.423)\n"
         "- Entertainment costs (2 CFR 200.438)\n"
         "- Lobbying or political activities (2 CFR 200.451)\n"
         "- Contributions or donations (2 CFR 200.434)\n"
         "- Fines and penalties (2 CFR 200.441)\n"
         "- Personal expenses not related to grant activities"),

        ("INDIRECT COSTS",
         "Indirect costs allowable at the negotiated rate of 15% of direct salaries "
         "and wages only. A current Negotiated Indirect Cost Rate Agreement (NICRA) "
         "must be on file with the agency."),

        ("DOCUMENTATION",
         "All expenditures must be supported by original receipts or equivalent "
         "documentation retained for three years after grant closeout. Purchases "
         "exceeding $75 require receipts per the IRS $75 rule."),
    ]
}

EXPENSES_1 = [
    # date, description, amount, category, vendor, notes
    ("2026-01-15", "Health educator salary — January", 2166.67, "personnel", "Payroll", "100% FTE month 1"),
    ("2026-01-15", "Program director salary — January", 1083.33, "personnel", "Payroll", "50% FTE month 1"),
    ("2026-01-20", "Fringe benefits — January payroll", 649.80, "fringe", "Payroll", "20% of salary"),
    ("2026-01-22", "Office supplies — paper, folders, pens", 87.50, "supplies", "Staples", "Receipt attached"),
    ("2026-01-28", "Mileage reimbursement — home visits", 54.60, "travel", "Staff", "42 mi x $1.30/mi"),
    ("2026-02-03", "Flight — regional conference Chicago", 342.00, "travel", "Southwest Airlines", "Economy class"),
    ("2026-02-04", "Hotel — 2 nights Chicago", 298.00, "travel", "Marriott Downtown", "GSA rate $149/night"),
    ("2026-02-04", "Meals — Chicago conference (3 days)", 177.00, "travel", "Per Diem", "GSA $59/day"),
    ("2026-02-10", "Nutrition counseling contract — Feb", 666.67, "contractual", "Healthy Living LLC", "Pre-approved"),
    ("2026-02-12", "Blood glucose test strips (200 units)", 380.00, "supplies", "MedSupply Co", "Program participants"),
    ("2026-02-15", "Health educator salary — February", 2166.67, "personnel", "Payroll", "100% FTE month 2"),
    ("2026-02-15", "Program director salary — February", 1083.33, "personnel", "Payroll", "50% FTE month 2"),
    ("2026-02-15", "Fringe benefits — February payroll", 649.80, "fringe", "Payroll", "20% of salary"),
    ("2026-03-05", "Laptop computer for health educator", 1249.00, "supplies", "Best Buy", "Essential for EHR access"),
    ("2026-03-08", "Printed patient education materials", 215.00, "supplies", "PrintQuick", "1,000 brochures"),
    ("2026-03-10", "Registration — Annual Public Health Conf.", 350.00, "training", "APHA", "Program Director"),
    ("2026-03-12", "Flight — Annual Public Health Conf. DC", 387.00, "travel", "Delta Airlines", "Economy class"),
    ("2026-03-13", "Hotel — 3 nights Washington DC", 447.00, "travel", "Hyatt", "GSA rate $149/night"),
    ("2026-03-13", "Meals — DC conference (4 days)", 236.00, "travel", "Per Diem", "GSA $59/day"),
    ("2026-03-15", "Health educator salary — March", 2166.67, "personnel", "Payroll", "100% FTE month 3"),
    ("2026-03-15", "Program director salary — March", 1083.33, "personnel", "Payroll", "50% FTE month 3"),
    ("2026-03-15", "Fringe benefits — March payroll", 649.80, "fringe", "Payroll", "20% of salary"),
    ("2026-03-20", "Indirect costs Q1 (15% direct salaries)", 1087.50, "indirect", "Admin", "Per NICRA"),
    ("2026-03-22", "Community outreach event — venue rental", 425.00, "contractual", "Community Center", "Pre-approved"),
]

# ─────────────────────────────────────────────────────────────────────────────
# EXAMPLE 2  Youth Education Foundation  ·  DoE Award #2026-DOE-207
# Mixed: some allowable, some flagged (alcohol, lobbying, personal, anomaly)
# ─────────────────────────────────────────────────────────────────────────────

GRANT_2 = {
    "title": "FEDERAL GRANT AGREEMENT",
    "grant_number": "2026-DOE-207",
    "grantee": "Youth Education Foundation",
    "agency": "U.S. Department of Education",
    "award_amount": "$120,000",
    "period": "July 1, 2026 - June 30, 2027",
    "cfda": "84.215",
    "sections": [
        ("PURPOSE",
         "Grant supports an after-school STEM tutoring program serving Title I "
         "elementary schools in an urban district. Program activities must comply "
         "with 2 CFR 200 and applicable Department of Education regulations."),

        ("BUDGET SUMMARY",
         "Personnel:             $72,000\n"
         "Fringe Benefits:       $14,400\n"
         "Travel:                 $5,500\n"
         "Supplies & Materials:   $8,000\n"
         "Technology:             $6,000\n"
         "Indirect Costs (15%):  $14,100\n"
         "Total:                $120,000"),

        ("ALLOWABLE COSTS",
         "Costs must satisfy 2 CFR 200.420: necessary, reasonable, allocable, and "
         "consistently applied. All costs must directly support STEM tutoring activities."),

        ("TRAVEL",
         "Travel costs allowable per 2 CFR 200.474. Economy class required for all "
         "air travel; upgrades are unallowable without prior agency approval. "
         "Per diem: meals $59/day (GSA); lodging not to exceed $175/night. "
         "Local mileage at current IRS rate. All travel must relate to program activities."),

        ("PERSONNEL",
         "Program Coordinator (100% FTE, $48,000/year): pre-approved.\n"
         "STEM Tutors (2 x 50% FTE, $12,000 each/year): pre-approved.\n"
         "Fringe benefit rate: 20% of salaries. Time sheets required."),

        ("SUPPLIES & TECHNOLOGY",
         "STEM curriculum materials, workbooks, and science kits are allowable. "
         "Technology purchases under $5,000 allowable if directly used in the program. "
         "No single technology item may exceed $1,500 without prior approval."),

        ("UNALLOWABLE COSTS",
         "Explicitly unallowable under this award:\n"
         "- Alcoholic beverages of any kind (2 CFR 200.423)\n"
         "- Entertainment not integral to the program (2 CFR 200.438)\n"
         "- Lobbying activities and political contributions (2 CFR 200.451)\n"
         "- Personal expenses unrelated to grant objectives (2 CFR 200.420)\n"
         "- First-class or business-class airfare without prior approval (2 CFR 200.474)\n"
         "- Contingency reserves or unspecified costs"),

        ("DOCUMENTATION",
         "Receipts and invoices required for all expenditures over $75. "
         "Electronic records acceptable when originals are unavailable. "
         "Records must be retained for five years after final expenditure report."),

        ("REPORTING",
         "Quarterly financial reports due 30 days after each quarter end. "
         "Final financial report and closeout package due 90 days after period end."),
    ]
}

EXPENSES_2 = [
    # date, description, amount, category, vendor, notes
    ("2026-07-01", "Program coordinator salary — July", 4000.00, "personnel", "Payroll", "100% FTE"),
    ("2026-07-01", "STEM tutor #1 salary — July", 1000.00, "personnel", "Payroll", "50% FTE"),
    ("2026-07-01", "STEM tutor #2 salary — July", 1000.00, "personnel", "Payroll", "50% FTE"),
    ("2026-07-01", "Fringe benefits — July payroll", 1200.00, "fringe", "Payroll", "20% of $6,000"),
    ("2026-07-10", "STEM curriculum kits — 30 students", 1350.00, "supplies", "Science4Kids", "Approved materials"),
    ("2026-07-15", "Laptop computers x3 for tutoring lab", 3600.00, "technology", "Dell", "$1,200 each"),
    ("2026-07-18", "Beer and wine — staff welcome party", 485.00, "food", "Total Beverage", "UNALLOWABLE"),
    ("2026-07-22", "Mileage — site visits to 4 schools", 78.00, "travel", "Staff", "60 mi x $1.30/mi"),
    ("2026-07-28", "Lobbying consultant — education advocacy", 3500.00, "professional", "DC Education Lobby", "UNALLOWABLE"),
    ("2026-08-01", "Program coordinator salary — August", 4000.00, "personnel", "Payroll", "100% FTE"),
    ("2026-08-01", "STEM tutor #1 salary — August", 1000.00, "personnel", "Payroll", "50% FTE"),
    ("2026-08-01", "STEM tutor #2 salary — August", 1000.00, "personnel", "Payroll", "50% FTE"),
    ("2026-08-01", "Fringe benefits — August payroll", 1200.00, "fringe", "Payroll", "20% of $6,000"),
    ("2026-08-05", "Science workbooks — 30 copies", 420.00, "supplies", "Scholastic", "Grade 3-5"),
    ("2026-08-12", "Personal vacation — family to Disney", 2800.00, "other", "Personal Charge", "UNALLOWABLE"),
    ("2026-08-14", "Flight to STEM conference — economy", 398.00, "travel", "American Airlines", "Receipt attached"),
    ("2026-08-15", "Hotel — 2 nights STEM conference", 350.00, "travel", "Hilton Garden Inn", "$175/night GSA"),
    ("2026-08-15", "Meals — conference (3 days)", 177.00, "travel", "Per Diem", "GSA $59/day"),
    ("2026-08-20", "Printer for tutoring center", 289.00, "technology", "HP", "Program use"),
    ("2026-09-01", "Program coordinator salary — September", 4000.00, "personnel", "Payroll", "100% FTE"),
    ("2026-09-01", "STEM tutor #1 salary — September", 1000.00, "personnel", "Payroll", "50% FTE"),
    ("2026-09-01", "STEM tutor #2 salary — September", 1000.00, "personnel", "Payroll", "50% FTE"),
    ("2026-09-01", "Fringe benefits — September payroll", 1200.00, "fringe", "Payroll", "20% of $6,000"),
    ("2026-09-10", "First-class flight to national conference", 2650.00, "travel", "Delta Airlines", "Requires approval"),
    ("2026-09-15", "Math manipulatives — classroom sets x5", 675.00, "supplies", "Nasco", "Pre-approved"),
    ("2026-09-20", "Indirect costs Q1 (15% direct salaries)", 3600.00, "indirect", "Admin", "Per NICRA"),
    ("2026-09-25", "Flight to STEM conference — economy", 398.00, "travel", "American Airlines", "Possible duplicate?"),
]


# ═════════════════════════════════════════════════════════════════════════════
# PDF GENERATOR
# ═════════════════════════════════════════════════════════════════════════════

def _styles():
    s = getSampleStyleSheet()
    h1 = ParagraphStyle("H1", parent=s["Heading1"],
                        fontSize=14, spaceAfter=4, textColor=colors.HexColor("#1a3560"))
    h2 = ParagraphStyle("H2", parent=s["Heading2"],
                        fontSize=11, spaceAfter=3, textColor=colors.HexColor("#1a3560"))
    body = ParagraphStyle("Body", parent=s["Normal"],
                          fontSize=9, leading=14, spaceAfter=4)
    pre = ParagraphStyle("Pre", parent=body, fontName="Courier", fontSize=8.5,
                         leftIndent=18, spaceAfter=2)
    return h1, h2, body, pre


def make_grant_pdf(grant: dict, path: str):
    doc = SimpleDocTemplate(path, pagesize=letter,
                            leftMargin=0.9*inch, rightMargin=0.9*inch,
                            topMargin=0.9*inch, bottomMargin=0.9*inch)
    h1, h2, body, pre = _styles()
    story = []

    story.append(Paragraph(grant["title"], h1))
    story.append(HRFlowable(width="100%", thickness=1.5,
                            color=colors.HexColor("#1a3560"), spaceAfter=8))

    meta = [
        ["Grant Number:", grant["grant_number"],
         "Grantee:", grant["grantee"]],
        ["Federal Agency:", grant["agency"],
         "CFDA Number:", grant["cfda"]],
        ["Award Amount:", grant["award_amount"],
         "Period of Performance:", grant["period"]],
    ]
    meta_table = Table(meta, colWidths=[1.3*inch, 2.2*inch, 1.5*inch, 2.2*inch])
    meta_table.setStyle(TableStyle([
        ("FONTNAME", (0,0), (-1,-1), "Helvetica"),
        ("FONTNAME", (0,0), (0,-1), "Helvetica-Bold"),
        ("FONTNAME", (2,0), (2,-1), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 9),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("BOTTOMPADDING", (0,0), (-1,-1), 5),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 14))

    for heading, text in grant["sections"]:
        story.append(Paragraph(f"SECTION — {heading}", h2))
        for line in text.split("\n"):
            line = line.strip()
            if not line:
                story.append(Spacer(1, 3))
            elif line.startswith("-"):
                story.append(Paragraph(line, pre))
            elif ":" in line and line.index(":") < 30:
                story.append(Paragraph(line, pre))
            else:
                story.append(Paragraph(line, body))
        story.append(Spacer(1, 8))

    doc.build(story)
    print(f"  [PDF] {os.path.basename(path)}")


def make_expense_pdf(expenses: list, grant: dict, path: str):
    doc = SimpleDocTemplate(path, pagesize=letter,
                            leftMargin=0.75*inch, rightMargin=0.75*inch,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)
    h1, h2, body, pre = _styles()
    story = []

    story.append(Paragraph("EXPENSE REPORT", h1))
    story.append(HRFlowable(width="100%", thickness=1.5,
                            color=colors.HexColor("#1a3560"), spaceAfter=8))

    meta = [
        ["Organization:", grant["grantee"], "Grant Number:", grant["grant_number"]],
        ["Report Period:", grant["period"], "Prepared:", "Finance Department"],
    ]
    mt = Table(meta, colWidths=[1.1*inch, 2.6*inch, 1.2*inch, 2.3*inch])
    mt.setStyle(TableStyle([
        ("FONTNAME", (0,0), (-1,-1), "Helvetica"),
        ("FONTNAME", (0,0), (0,-1), "Helvetica-Bold"),
        ("FONTNAME", (2,0), (2,-1), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 9),
        ("BOTTOMPADDING", (0,0), (-1,-1), 5),
    ]))
    story.append(mt)
    story.append(Spacer(1, 14))

    header = ["#", "Date", "Description", "Amount", "Category", "Vendor"]
    col_w = [0.3*inch, 0.85*inch, 2.9*inch, 0.75*inch, 0.85*inch, 1.3*inch]
    rows = [header]
    total = 0.0
    for i, (date, desc, amt, cat, vendor, _notes) in enumerate(expenses, 1):
        rows.append([str(i), date, desc, f"${amt:,.2f}", cat, vendor])
        total += amt
    rows.append(["", "", "TOTAL", f"${total:,.2f}", "", ""])

    tbl = Table(rows, colWidths=col_w, repeatRows=1)
    grey = colors.HexColor("#1a3560")
    light = colors.HexColor("#dce6f4")
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), grey),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 8),
        ("ALIGN", (3,0), (3,-1), "RIGHT"),
        ("ALIGN", (0,0), (1,-1), "CENTER"),
        ("ROWBACKGROUNDS", (0,1), (-1,-2), [colors.white, light]),
        ("FONTNAME", (0,-1), (-1,-1), "Helvetica-Bold"),
        ("LINEBELOW", (0,-2), (-1,-2), 1, grey),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
        ("TOPPADDING", (0,0), (-1,-1), 4),
        ("GRID", (0,0), (-1,-1), 0.3, colors.HexColor("#b0c4de")),
    ]))
    story.append(tbl)
    doc.build(story)
    print(f"  [PDF] {os.path.basename(path)}")


# ═════════════════════════════════════════════════════════════════════════════
# EXCEL GENERATOR
# ═════════════════════════════════════════════════════════════════════════════

NAVY = "1A3560"
LIGHT_BLUE = "DCE6F4"
YELLOW = "FFF2CC"
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", bold=True, size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
THIN = Border(
    left=Side(style="thin", color="B0C4DE"),
    right=Side(style="thin", color="B0C4DE"),
    top=Side(style="thin", color="B0C4DE"),
    bottom=Side(style="thin", color="B0C4DE"),
)


def _set(ws, row, col, value, font=None, fill=None, align=None, border=None, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:      cell.font = font
    if fill:      cell.fill = fill
    if align:     cell.alignment = align
    if border:    cell.border = border
    if number_format: cell.number_format = number_format
    return cell


def make_grant_excel(grant: dict, path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Grant Agreement"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 60

    navy_fill = PatternFill("solid", fgColor=NAVY)
    blue_fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    white_fill = PatternFill("solid", fgColor="FFFFFF")

    # Title
    ws.merge_cells("A1:B1")
    _set(ws, 1, 1, grant["title"],
         font=Font("Calibri", bold=True, size=14, color="FFFFFF"),
         fill=navy_fill, align=CENTER)
    ws.row_dimensions[1].height = 28

    # Metadata block
    meta_rows = [
        ("Grant Number", grant["grant_number"]),
        ("Grantee Organization", grant["grantee"]),
        ("Federal Agency", grant["agency"]),
        ("CFDA Number", grant["cfda"]),
        ("Award Amount", grant["award_amount"]),
        ("Period of Performance", grant["period"]),
    ]
    for i, (label, val) in enumerate(meta_rows, 2):
        _set(ws, i, 1, label, font=BOLD_FONT,
             fill=blue_fill, align=LEFT, border=THIN)
        _set(ws, i, 2, val, font=BODY_FONT,
             fill=white_fill, align=LEFT, border=THIN)

    r = len(meta_rows) + 3
    for heading, text in grant["sections"]:
        ws.merge_cells(f"A{r}:B{r}")
        _set(ws, r, 1, heading,
             font=Font("Calibri", bold=True, size=11, color="FFFFFF"),
             fill=navy_fill, align=LEFT)
        ws.row_dimensions[r].height = 20
        r += 1
        for line in text.split("\n"):
            ws.merge_cells(f"A{r}:B{r}")
            _set(ws, r, 1, line.strip() if line.strip() else "",
                 font=BODY_FONT, fill=white_fill, align=LEFT, border=THIN)
            ws.row_dimensions[r].height = 16 if line.strip() else 8
            r += 1
        r += 1

    wb.save(path)
    print(f"  [XLS] {os.path.basename(path)}")


def make_expense_excel(expenses: list, grant: dict, path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expense Report"

    navy_fill = PatternFill("solid", fgColor=NAVY)
    blue_fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    yellow_fill = PatternFill("solid", fgColor=YELLOW)
    white_fill = PatternFill("solid", fgColor="FFFFFF")

    # Widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 44
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 26

    # Title row
    ws.merge_cells("A1:G1")
    _set(ws, 1, 1, "EXPENSE REPORT",
         font=Font("Calibri", bold=True, size=14, color="FFFFFF"),
         fill=navy_fill, align=CENTER)
    ws.row_dimensions[1].height = 28

    # Meta block
    ws.merge_cells("A2:G2")
    _set(ws, 2, 1,
         f"Organization: {grant['grantee']}    |    Grant #: {grant['grant_number']}    |    Period: {grant['period']}",
         font=Font("Calibri", bold=True, size=10),
         fill=blue_fill, align=CENTER)
    ws.row_dimensions[2].height = 18

    # Header row
    headers = ["#", "Date", "Description", "Amount", "Category", "Vendor", "Notes"]
    for col, h in enumerate(headers, 1):
        _set(ws, 3, col, h, font=HEADER_FONT,
             fill=navy_fill, align=CENTER, border=THIN)
    ws.row_dimensions[3].height = 20

    # Data rows
    total = 0.0
    for i, (date, desc, amt, cat, vendor, notes) in enumerate(expenses, 1):
        r = i + 3
        fill = blue_fill if i % 2 == 0 else white_fill
        _set(ws, r, 1, i,        font=BODY_FONT, fill=fill, align=CENTER, border=THIN)
        _set(ws, r, 2, date,     font=BODY_FONT, fill=fill, align=CENTER, border=THIN)
        _set(ws, r, 3, desc,     font=BODY_FONT, fill=fill, align=LEFT,   border=THIN)
        cell_amt = _set(ws, r, 4, amt, font=BODY_FONT, fill=fill, align=RIGHT, border=THIN,
                        number_format='"$"#,##0.00')
        _set(ws, r, 5, cat,      font=BODY_FONT, fill=fill, align=CENTER, border=THIN)
        _set(ws, r, 6, vendor,   font=BODY_FONT, fill=fill, align=LEFT,   border=THIN)
        _set(ws, r, 7, notes,    font=BODY_FONT, fill=fill, align=LEFT,   border=THIN)
        # Highlight UNALLOWABLE notes
        if "UNALLOWABLE" in notes:
            for col in range(1, 8):
                ws.cell(row=r, column=col).fill = yellow_fill
        total += amt
        ws.row_dimensions[r].height = 16

    # Totals row
    tr = len(expenses) + 4
    ws.merge_cells(f"A{tr}:C{tr}")
    _set(ws, tr, 1, "TOTAL", font=BOLD_FONT, fill=navy_fill,
         align=RIGHT, border=THIN)
    ws.cell(row=tr, column=1).font = Font("Calibri", bold=True, size=10, color="FFFFFF")
    _set(ws, tr, 4, total, font=Font("Calibri", bold=True, size=10, color="FFFFFF"),
         fill=navy_fill, align=RIGHT, border=THIN, number_format='"$"#,##0.00')
    for col in [5, 6, 7]:
        _set(ws, tr, col, "", fill=navy_fill, border=THIN)
    ws.row_dimensions[tr].height = 18

    # Freeze header
    ws.freeze_panes = "A4"

    # Summary sheet
    ws2 = wb.create_sheet("Category Summary")
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 14

    ws2.merge_cells("A1:C1")
    _set(ws2, 1, 1, "SPENDING BY CATEGORY",
         font=Font("Calibri", bold=True, size=12, color="FFFFFF"),
         fill=navy_fill, align=CENTER)
    ws2.row_dimensions[1].height = 22

    for col, h in enumerate(["Category", "Total Spent", "# Items"], 1):
        _set(ws2, 2, col, h, font=HEADER_FONT, fill=navy_fill, align=CENTER, border=THIN)

    cat_totals = {}
    for _, _, amt, cat, _, _ in expenses:
        cat_totals[cat] = cat_totals.get(cat, (0.0, 0))
        cat_totals[cat] = (cat_totals[cat][0] + amt, cat_totals[cat][1] + 1)

    for r, (cat, (spent, count)) in enumerate(sorted(cat_totals.items()), 3):
        fill = blue_fill if r % 2 == 0 else white_fill
        _set(ws2, r, 1, cat,   font=BODY_FONT, fill=fill, align=LEFT,   border=THIN)
        _set(ws2, r, 2, spent, font=BODY_FONT, fill=fill, align=RIGHT,  border=THIN, number_format='"$"#,##0.00')
        _set(ws2, r, 3, count, font=BODY_FONT, fill=fill, align=CENTER, border=THIN)

    wb.save(path)
    print(f"  [XLS] {os.path.basename(path)}")


# ═════════════════════════════════════════════════════════════════════════════
# MAIN
# ═════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    print(f"\nGenerating sample documents in: {OUT}\n")

    print("Example 1 — Community Health Initiative (HHS, clean expenses):")
    make_grant_pdf(GRANT_1,   os.path.join(OUT, "example1_grant_agreement_HHS101.pdf"))
    make_grant_excel(GRANT_1, os.path.join(OUT, "example1_grant_agreement_HHS101.xlsx"))
    make_expense_pdf(EXPENSES_1,   GRANT_1, os.path.join(OUT, "example1_expense_report_HHS101.pdf"))
    make_expense_excel(EXPENSES_1, GRANT_1, os.path.join(OUT, "example1_expense_report_HHS101.xlsx"))

    print("\nExample 2 — Youth Education Foundation (DoE, mixed allowable/flagged):")
    make_grant_pdf(GRANT_2,   os.path.join(OUT, "example2_grant_agreement_DOE207.pdf"))
    make_grant_excel(GRANT_2, os.path.join(OUT, "example2_grant_agreement_DOE207.xlsx"))
    make_expense_pdf(EXPENSES_2,   GRANT_2, os.path.join(OUT, "example2_expense_report_DOE207.pdf"))
    make_expense_excel(EXPENSES_2, GRANT_2, os.path.join(OUT, "example2_expense_report_DOE207.xlsx"))

    print(f"\nDone. 8 files written to data/sample_documents/")
